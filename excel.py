import math
#讀取檔案
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook(filename = '110年零用金流向0610.xlsx', data_only=True)
ws = wb['work']
wb2 = load_workbook(filename = 'work_sample.xlsx', data_only=True)
ws2 = wb2['sheet']

#簽證號
def wordfinder_number(searchString):
    j = 1
    for i in range(1, ws.max_row):
        if searchString == ws.cell(i,j).value and ws.cell(i,j + 6).value == '物品':
            ws2['U1'].value = '335-'
            ws2['C6'].value = '消防業務－各救災救護大隊－業務費－物品'
            ws2['P6'].value = ws.cell(i,j + 7).value + '\n代墊人: '                   
        elif searchString == ws.cell(i,j).value and ws.cell(i,j + 6).value == '一般':
            ws2['U1'].value = '396-'
            ws2['C6'].value = '消防業務－各救災救護大隊－業務費－一般事務費'
            ws2['P6'].value = ws.cell(i,j + 7).value + '\n代墊人: '        
        elif searchString == ws.cell(i,j).value and ws.cell(i,j + 6).value == '電費':
            ws2['U1'].value = '213-'
            ws2['C6'].value = '消防業務－各救災救護大隊－業務費－水電費'
            ws2['P6'].value = ws.cell(i,j + 7).value + '\n代墊人: '        
        elif searchString == ws.cell(i,j).value and ws.cell(i,j + 6).value == '水費':
            ws2['U1'].value = '158-'
            ws2['C6'].value = '消防業務－各救災救護大隊－業務費－水費'
            ws2['P6'].value = ws.cell(i,j + 7).value + '\n代墊人: '        
        elif searchString == ws.cell(i,j).value and ws.cell(i,j + 6).value == '電話':
            ws2['U1'].value = '278-'
            ws2['C6'].value = '消防業務－各救災救護大隊－業務費－通訊費'
            ws2['P6'].value = ws.cell(i,j + 7).value + '\n代墊人: '

# 依照項次加入品名
list_name = []
def wordfinder_name(searchString):
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            if searchString == ws.cell(i,j).value:
                list_name.append(ws.cell(i,j + 2).value)
                ws2.cell(len(list_name) + 20, 1).value = ws.cell(i,j + 2).value 

# 依照項次加入單價
list_price = []
def wordfinder(searchString):
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            if searchString == ws.cell(i,j).value:
                list_price.append(ws.cell(i,j + 5).value)
                ws2.cell(len(list_price) + 20, 13).value = ws.cell(i,j + 5).value

# 依照項次加入數量
list_quantity = []
def wordfinder_quantity(searchString):
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            if searchString == ws.cell(i,j).value:
                list_quantity.append(ws.cell(i,j + 4).value)
                ws2.cell(len(list_quantity) + 20, 10).value = ws.cell(i,j + 4).value

# 依照項次加入單位
list_type = []
def wordfinder_type(searchString):
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            if searchString == ws.cell(i,j).value:
                list_type.append(ws.cell(i,j + 3).value)
                ws2.cell(len(list_type) + 20, 7).value = ws.cell(i,j + 3).value

# 加入廠商名稱
list_vender = []
def wordfinder_vender(searchString):
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            if searchString == ws.cell(i,j).value:
                list_vender.append(ws.cell(i,j + 1).value)

'''type_item = []
with open('type.csv', 'r', encoding = 'utf8') as f:
    for d in f:
        d = d.replace('\n', '')
        type_item.append(d)
        print(d)

while True:
    input_item = input('是否製作請示單: (y/n)')'''





d = input('請問這次要印的請示單是: ')
type_list = []
type_list.append(d)
wordfinder(d)
wordfinder_quantity(d)
wordfinder_name(d)
wordfinder_type(d)
wordfinder_vender(d)
wordfinder_number(d)

#單價乘數量
total_price = [x*y for x,y in zip(list_quantity,list_price)]
n = 0
for i in total_price:
    ws2.cell(n + 21, 17).value = round(i) 
    n += 1

#金額總和
listSum = sum(total_price)
listSum = round(listSum)
print(listSum)
ws2.cell(27, 17).value = str(listSum)

#下方欄
ws2.cell(29, 1).value = '2. ■本案經詢價擬以' + str(ws2.cell(27, 17).value) + '元交由 ' + list_vender[0] +' 辦理，並經驗收合格後付款。' 

#清單金額填到請示單

mon = listSum // 10000
senn = (listSum - mon * 10000) // 1000
hyaku = (listSum - mon * 10000- senn * 1000) // 100
ju = (listSum - mon * 10000 - senn * 1000 - hyaku * 100) // 10
enn = listSum % 10

if mon > 0: #萬
    ws2['K6'].value = mon
elif mon == 0:
    ws2['K6'].value = None

if senn > 0: #千
    ws2['L6'].value = senn
elif senn == 0 and mon > 0:
    ws2['L6'].value = 0
elif senn == 0:
    ws2['L6'].value = None

if hyaku > 0: #百
    ws2['M6'].value = hyaku
elif hyaku == 0 and mon > 0:
    ws2['M6'].value = 0
elif hyaku == 0 and senn > 0:
    ws2['M6'].value = 0
elif hyaku == 0:
    ws2['M6'].value = None

if ju > 0: #十
    ws2['N6'].value = ju
elif ju == 0 and mon > 0:
    ws2['N6'].value = 0
elif ju == 0 and senn > 0:
    ws2['N6'].value = 0
elif ju == 0 and hyaku > 0:
    ws2['N6'].value = 0
elif ju == 0:
    ws2['N6'].value = None
ws2['O6'].value = enn

#存檔
def wordfinder_save(searchString):
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            if searchString == ws.cell(i,j).value:
                wb2.save(filename = searchString +'.xlsx')

bian = Side(style='thin', color='000000') 
border = Border(top=bian, bottom=bian, left=bian, right=bian, diagonal=bian, diagonalDown=True)
ws2['B13'].border = border
ws2['B15'].border = border
ws2['B33'].border = border
ws2['B34'].border = border
wordfinder_save(d)